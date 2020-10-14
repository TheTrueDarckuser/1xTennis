from time import sleep
from selenium import webdriver, common
from selenium.webdriver.common.keys import Keys
from selenium.webdriver import ActionChains
from urllib3 import exceptions
import json
import time
import datetime
import requests
import winsound
import openpyxl

useAPI = False
apiLiveIP = 'http://b-api.com/1x-api/live/Tennis/'
oneXbetUrl = 'https://1x-bet93029.com/ru/live/Tennis/'

browser = webdriver.Firefox()
browser.set_window_size(2500, 2000)


class OneXStavka(object):
    def __init__(self, browser, apiLiveIP, useAPI):
        self.browser = browser
        self.apiLiveIP = apiLiveIP
        self.useAPI = useAPI

        try:
            file = open("buff.json", 'x')
            file.close()
        except FileExistsError:
            pass

        try:
            file = open("statistic.json", 'x')
            file.close()
        except FileExistsError:
            pass


        try:
            self.excel_document = openpyxl.load_workbook('information.xlsx')
        except FileNotFoundError:
            filepath = "information.xlsx"
            wb = openpyxl.Workbook()
            wb.create_sheet('Разница 7')

            def setTitle(activeSheet):
                titleData = ['Лига', 'Пары', 'Женщины', 'М+Ж', 'Имена', 'Счет', 'Поб 1', 'Поб 1 нов', 'Поб 2',
                             'Поб 2 нов', 'Фаворит', 'Тотал Б',
                             'Тотал Б нов', 'Тотал', 'Тотал нов', 'Тотал М', 'Тотал М нов',
                             'ИД 1 Б нов',
                             'ИД 1 нов',
                             'ИД 1 М нов',
                             'ИД 2 Б нов',
                             'ИД 2 нов',
                             'ИД 2 М нов',
                             'Ф 1', 'Ф 1 нов',
                             'Ф', 'Ф нов',
                             'Ф 2', 'Ф 2 нов',
                             'Результат', 'Ссылка']
                for row in range(len(titleData)):
                    activeCell = activeSheet.cell(1, row + 1)
                    activeCell.value = titleData[row]

            wb.create_sheet('Разница 7,5')
            wb.create_sheet('Разница 8')
            wb.create_sheet('Разница 8,5')
            wb.create_sheet('Разница 9')
            wb.remove(wb[wb.sheetnames[0]])
            for tab in wb.sheetnames:
                setTitle(wb[tab])
            wb.save(filepath)
            self.excel_document = openpyxl.load_workbook('information.xlsx')

    def readInfo(self, file):
        with open(file, 'r') as file:
            try:
                oldInformation = json.load(file)
            except:
                oldInformation = {}
        return oldInformation

    def writeInfo(self, info, nameFile):
        with open(nameFile, 'w') as file:
            json.dump(info, file)

    def apiLive(self):
        return requests.get(self.apiLiveIP).json()

    def algoritm(self):
        self.browser.get(oneXbetUrl)
        sleep(5)

        countForReloadPage = 0
        while 1:
            countForReloadPage += 1
            if countForReloadPage > 200:
                countForReloadPage = 0
                browser.refresh()
                time.sleep(5)

            try:
                games = browser.find_element_by_id('games_content').find_element_by_xpath('div/div/div[1]')
                games = games.find_elements_by_xpath('div')

                for game in games:
                    try:
                        game = game.find_elements_by_xpath('div')
                    except common.exceptions.StaleElementReferenceException:
                        print('103 common.exceptions.StaleElementReferenceException:')
                        continue

                    liga = game[0].text.split('\n')[0]
                    oldURLS = self.readInfo('buff.json')
                    oldURLS = list(oldURLS.values())
                    oldURLS = [i['url'] for i in oldURLS]

                    matches = game[1:]
                    for match in matches:
                        try:
                            name = match.find_element_by_class_name('c-events__name').text
                            while name.find('(') != -1:
                                name = name[:name.find('(')] + name[name.find(')') + 1:]
                            name = name.split('\n')
                            name = [i.strip() for i in name]
                            name = '\n'.join(name)

                            total = match.find_elements_by_class_name('c-bets__bet')[4].text
                            if total == '-':
                                continue

                            firstKoef = match.find_elements_by_class_name('c-bets__bet')[0].text
                            if firstKoef == '-':
                                continue
                            secondKoef = match.find_elements_by_class_name('c-bets__bet')[2].text
                            if secondKoef == '-':
                                continue

                            totalL = match.find_elements_by_class_name('c-bets__bet')[5].text
                            totalM = match.find_elements_by_class_name('c-bets__bet')[3].text
                            idT1L = match.find_elements_by_class_name('c-bets__bet')[11].text
                            idT1 = match.find_elements_by_class_name('c-bets__bet')[10].text
                            idT1M = match.find_elements_by_class_name('c-bets__bet')[9].text
                            idT2L = match.find_elements_by_class_name('c-bets__bet')[14].text
                            idT2 = match.find_elements_by_class_name('c-bets__bet')[13].text
                            idT2M = match.find_elements_by_class_name('c-bets__bet')[12].text
                            odds1 = match.find_elements_by_class_name('c-bets__bet')[6].text
                            odds = match.find_elements_by_class_name('c-bets__bet')[7].text
                            odds2 = match.find_elements_by_class_name('c-bets__bet')[8].text
                            url = match.find_element_by_class_name('c-events__name').get_attribute('href')

                            score = match.find_elements_by_class_name('c-events-scoreboard__cell--all')
                            score = [i.text for i in score]

                        except common.exceptions.StaleElementReferenceException:
                            print('150 common.exceptions.StaleElementReferenceException')
                            continue

                        if self.useAPI:
                            try:
                                infOld = self.apiLive()
                                oldFirstKoef = float(infOld[liga][name]['firstKoef'])
                                oldSecondKoef = float(infOld[liga][name]['secondKoef'])
                                oldTotal = float(infOld[liga][name]['total'])
                            except (KeyError, ValueError):
                                continue
                            except requests.exceptions.ConnectionError:
                                print('Ошибка соединения с API')
                                continue
                        else:
                            try:
                                infOld = self.readInfo("lineTennisExpand.json")
                                oldFirstKoef = float(infOld[liga][name]['firstKoef'])
                                oldSecondKoef = float(infOld[liga][name]['secondKoef'])
                                oldTotal = float(infOld[liga][name]['total'])
                            except (KeyError, ValueError):
                                continue

                        if oldFirstKoef <= oldSecondKoef:
                            favoriteNumber = "1"
                            favorite = oldFirstKoef
                            outsider = oldSecondKoef
                        else:
                            favoriteNumber = "2"
                            favorite = oldSecondKoef
                            outsider = oldFirstKoef

                        print(liga)
                        print(name)
                        print(total)
                        print('Фаворит коэф', favorite)
                        print('Разница тоталов', (float(total) - float(oldTotal)))
                        print(url)
                        print('\n\n')

                        if (float(total) - float(oldTotal)) >= 7:
                            self.addInfoJson(liga, name, [oldFirstKoef,firstKoef], [oldSecondKoef, secondKoef],
                                             [infOld[liga][name]['totalL'], float(totalL)], [infOld[liga][name]['total'], float(total)], [infOld[liga][name]['totalM'], float(totalM)],
                                             [infOld[liga][name]['idT1L'], float(idT1L)], [infOld[liga][name]['idT1'], float(idT1)], [infOld[liga][name]['idT1M'], float(idT1M)],
                                             [infOld[liga][name]['idT2L'], float(idT2L)], [infOld[liga][name]['idT2'], float(idT2)], [infOld[liga][name]['idT2M'], float(idT2M)],
                                             [infOld[liga][name]['odds1'], odds1], [infOld[liga][name]['odds'], odds], [infOld[liga][name]['odds2'], odds2],
                                             url, score, infOld[liga][name]['timeStart'])

                            # print(liga)
                            # print(name)
                            # print(total)
                            # print(url)
                            # print('\n\n')



            except (exceptions.ProtocolError, common.exceptions.NoSuchElementException,common.exceptions.StaleElementReferenceException):
                print('213 exceptions.ProtocolError, common.exceptions.NoSuchElementException,common.exceptions.StaleElementReferenceException')
                browser.refresh()
                time.sleep(5)


    def addInfoJson(self,liga, name, win1, win2, totalL, total, totalM, idT1L, idT1, idT1M, idT2L, idT2, idT2M, odds1, odds, odds2, url, score, timeStart):
        women = False
        pairs = False
        mixed = False
        if liga.lower().find('женщины') != -1:
            women = True
        if liga.lower().find('пары') != -1:
            pairs = True
        if liga.lower().find('mixed') != -1:
            mixed = True
        name1, name2 = name.split('\n')
        # Тут анализируем полученную информацию, разбиваем имена, регистрируем время, приводим к нужным типам переменных
        # Решть как передавать значения разницы тоталов: в главном ключе и значение передавать потом, либо передавать в контексте
        # Возможна ошибка если ключа нет с разницей тоталов, то нужно создавать
        #Доделать словарь
        Dict = self.readInfo('statistic.json')
        if len(Dict[float(float(total[0]) - float(total[1]))][url]) == 0:
            Dict.update({ float(float(total[0]) - float(total[1])) : {url: {'liga' : liga, 'name1': name1.strip(),
                                                                                     'name2': name2.strip(), 'women':women,
                                                                                     'pairs':pairs, 'mixed':mixed,
                                                                                     'win1':win1,'win2':win2,
                                                                                     'totalL':totalL, 'total':total, 'totalM':totalM,
                                                                                     'idT1L':idT1L, 'idT1':idT1, 'idT1M':idT1M,
                                                                                     'idT2L':idT2L, 'idT2':idT2, 'idT1M':idT2M,
                                                                                     'odds1':odds1, 'odds':odds, 'odds2':odds2,
                                                                                     'url':url, 'endTotal':None, 'endScore':None, 'result':None, 'score':score, 'timeStart':timeStart, }}})
        self.writeInfo(Dict, 'statistic.json')


    def addInfo(self, workSheetNum, infOld, liga, name, score, firstKoef, secondKoef, favorite, totalMore, total, totalLess, idTotalFirstMore, idTotalFirst, idTotalFirstLess, idTotalSecondMore, idTotalSecond, idTotalSecondLess, oddsFirst, odds, oddsSecond, url):

        women = 0
        pairs = 0
        mixed = 0
        if liga.lower().find('женщины') != -1:
            women = 1
        if liga.lower().find('пары') != -1:
            pairs = 1
        if liga.lower().find('mixed') != -1:
            mixed = 1



        line = [liga, pairs, women, mixed, name, ' '.join(score), infOld[liga][name]['firstKoef'], firstKoef,
                infOld[liga][name]['secondKoef'], secondKoef, favorite, infOld[liga][name]['totalMore'], totalMore,
                infOld[liga][name]['total'], total, infOld[liga][name]['totalLess'], totalLess,
                idTotalFirstMore, idTotalFirst, idTotalFirstLess, idTotalSecondMore, idTotalSecond,
                idTotalSecondLess, infOld[liga][name]['oddsFirst'], oddsFirst, infOld[liga][name]['odds'], odds,
                infOld[liga][name]['oddsSecond'], oddsSecond,
                '-*-', url]

        selectedSheet = self.excel_document[self.excel_document.sheetnames[workSheetNum - 1]]
        selectedSheet.append(line)
        self.excel_document.save('information.xlsx')


controlPanel = OneXStavka(browser, apiLiveIP, useAPI)

while 1:
    # try:
    controlPanel.algoritm()
    # except:
    #     print('Полный перезапуск автоставки')
    #     try:
    #         browser.quit()
    #         browser.quit()
    #     except:
    #         pass
    #     browser = webdriver.Firefox()
    #     print('Запуск браузера')
    #     controlPanel = OneXStavka(browser, url, login, password, amount, apiLiveIP, useAPI)
